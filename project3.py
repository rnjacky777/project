import openpyxl,json,time
import os
if not os.path.exists('account_book.xlsx'):
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    year = time.localtime().tm_year
    Month = time.localtime().tm_mon
    sheet.title = str(year)+'_'+str(Month)
    wb.save(r'account_book.xlsx')
wb = openpyxl.load_workbook('account_book.xlsx')
year = time.localtime().tm_year
Month = time.localtime().tm_mon
sheetname = str(year)+'_'+str(Month)
#wb.sheetnames
sheet = wb[sheetname]
#sheet['A1'].value = 1
#sheet['A1'].value
#wb.save('account_book.xlsx')
i = 1
while True:
    if sheet['A'+str(i)].value == None:
       break
    else:
        i+=1
print(i)
    
def addprice(name,price,date):
    global i
    print(i)
    sheet['A'+str(i)].value = date
    sheet['B'+str(i)].value = name
    sheet['C'+str(i)].value = price
    i += 1

while True:
    
    print('1.記帳 2.離開')
    try:
        a = int(input())
    except:
        continue
    if a == 1:
        date = input('請輸入日期')
        name = input('請輸入消費名稱')
        price = input('請輸入消費價格')
        addprice(name,price,date)
    elif a == 2:
        wb.save('account_book.xlsx')
        break
        
    

